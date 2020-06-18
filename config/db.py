import sys
from flask import jsonify
import pymysql
import json
from collections import OrderedDict
from pymysql.cursors import DictCursorMixin, Cursor
import openpyxl
from unipath import Path
from datetime import date


def sql_fetch_json(cursor):
    keys = []
    for column in cursor.description:
        keys.append(column[0])
    key_number = len(keys)

    json_data = []
    for row in cursor.fetchall():
        item = dict()
        for q in range(key_number):
            item[keys[q]] = row[q]
        json_data.append(item)
    print json_data
    return json_data

def modificar_Excel_hojas(sql,a,b,c,d,ruta_file,hoja_index,fil_ini,col_ini,doc):
        try:
            print sql
            hoja = doc.get_sheet_by_name(doc.sheetnames[hoja_index])
            db = pymysql.connect(a,b,c,d)
            cursor = db.cursor(OrderedDictCursor)
            cursor.execute(sql)
            cabecera=cursor.description
            result = cursor.fetchall()
            temp_col=col_ini
            temp_fil=fil_ini
            # for x in cabecera:
            #     if(fil_ini==fil_ini):
            #         hoja.cell(row=fil_ini, column=col_ini).value = x[0] 
            #         col_ini= col_ini+1
            # fil_ini=fil_ini+1 
            for y in result:
                    col_ini=temp_col
                    for z in cabecera:
                        hoja.cell(row=fil_ini, column=col_ini).value =y[str(z[0])]
                        col_ini= col_ini+1
                    fil_ini=fil_ini+1
            cursor.close()
            db.close()
             
        except db.Error as error:
                print "Unexpected error:{0}", error
        
        finally:
                print("MySQL connection is closed")
class OrderedDictCursor(DictCursorMixin, Cursor):
    dict_type = OrderedDict

class Metodos():


    
    def set_list(self,sql,a,b,c,d):
        try:
            ############### CONFIGURAR ESTO ###################
            # Abre conexion con la base de datos
            db = pymysql.connect(a,b,c,d)
            ##################################################

            # prepare a cursor object using cursor() method
            cursor = db.cursor(OrderedDictCursor)

            # ejecuta el SQL query usando el metodo execute().
            cursor.execute(sql)
           
            result = cursor.fetchall()
               
           
            # desconecta del servidor
            cursor.close()
            db.close()
            
            if result != None :
                return jsonify({"type":"success","status":200,"data":result})
        except db.Error as error:
                print "Unexpected error:{0}", error
        
        finally:
                print("MySQL connection is closed")

    def set_insert(self,sql,a,b,c,d):
        try:
                db = pymysql.connect(a,b,c,d)
                cursor = db.cursor()
                cursor.execute(sql)
                db.commit()
                # cursor.rowcount cuando sea update
                data=cursor.lastrowid
                cursor.close()
                db.close()
                print(data, "Record inserted successfully into Laptop table")
                if data != None :
                     return jsonify({"type":"success","status":200,"data":data})
                
        except db.Error as error:
                print "Unexpected error:{0}", error
        finally:
                print("MySQL connection is closed")



    def modificar_Excel(self,sql,a,b,c,d,ruta_file):
        try:
            ############### CONFIGURAR ESTO ###################
            # Abre conexion con la base de datos
            today = date.today()
            print(today)
            db = pymysql.connect(a,b,c,d)
            # prepare a cursor object using cursor() method
            cursor = db.cursor(OrderedDictCursor)
            # ejecuta el SQL query usando el metodo execute().
            cursor.execute(sql)
            # cabecera=cursor.description
            result = cursor.fetchall()
            doc = openpyxl.load_workbook('./archivos/'+ruta_file)
            for data in result:    
                fil_ini=data['fil']
                col_ini=data['col']
                if data['nivel2'] != 1:
                    result=modificar_Excel_hojas(data['query'],data['host'],data['user'],data['password'],data['db'],ruta_file, data['hoja'] - 1,fil_ini,col_ini,doc)
                else:
                    print data['nivel2']
            doc.active=0
            doc.save('./reportes/'+str(today)+ruta_file)              
            cursor.close()
            db.close()
        except db.Error as error:
                print "Unexpected error:{0}", error
        
        finally:
                print("MySQL connection is closed")
    
  


metodos=Metodos()


